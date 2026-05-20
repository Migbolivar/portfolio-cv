"""
Proyecto 1 - Excel: Análisis de Ventas Mensuales
================================================
Objetivo: Crear tabla dinámica de ventas por cliente y producto,
          con automatización mediante pandas + openpyxl.

Uso:
    python generar_excel.py

Requisitos: pandas, openpyxl
    pip install pandas openpyxl
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from pathlib import Path

# ── Cargar datos ──────────────────────────────────────────
DATA_DIR = Path(__file__).parent.parent / 'datos'
OUTPUT = Path(__file__).parent / 'ventas_reporte.xlsx'

df = pd.read_csv(DATA_DIR / 'ventas.csv', parse_dates=['Fecha'])
df['Mes'] = df['Fecha'].dt.to_period('M').astype(str)
df['Total'] = df['Cantidad'] * df['PrecioUnitario']
df['Año'] = df['Fecha'].dt.year

print(f'✓ Datos cargados: {len(df)} registros')
print(f'  Período: {df["Fecha"].min().date()} → {df["Fecha"].max().date()}')
print(f'  Ventas totales: ${df["Total"].sum():,.2f}')

# ── Crear workbook ────────────────────────────────────────
wb = Workbook()

# ────── HOJA 1: Datos crudos ──────────────────────────────
ws1 = wb.active
ws1.title = 'Datos'
for r in dataframe_to_rows(df.drop(columns=['Mes', 'Total', 'Año']), index=False, header=True):
    ws1.append(r)

# Formato de encabezados
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
for cell in ws1[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

# ────── HOJA 2: Tabla dinámica (ventas por Cliente × Producto) ──
ws2 = wb.create_sheet('Pivot Cliente-Producto')
pivot = df.pivot_table(
    values='Total',
    index='Cliente',
    columns='Producto',
    aggfunc='sum',
    fill_value=0
)
pivot['Total Cliente'] = pivot.sum(axis=1)
pivot = pivot.sort_values('Total Cliente', ascending=False)

for r in dataframe_to_rows(pivot.reset_index(), index=False, header=True):
    ws2.append(r)

for cell in ws2[1]:
    cell.font = header_font
    cell.fill = header_fill

# Formato moneda
money_fmt = '$#,##0'
for row in ws2.iter_rows(min_row=2, min_col=2, max_col=len(pivot.columns)+2):
    for cell in row:
        if isinstance(cell.value, (int, float)):
            cell.number_format = money_fmt

# ────── HOJA 3: Ventas mensuales ──────────────────────────
ws3 = wb.create_sheet('Ventas Mensuales')
mensual = df.groupby('Mes')['Total'].sum().reset_index()
mensual['Acumulado'] = mensual['Total'].cumsum()

for r in dataframe_to_rows(mensual, index=False, header=True):
    ws3.append(r)

for cell in ws3[1]:
    cell.font = header_font
    cell.fill = header_fill

# Gráfico de barras
chart = BarChart()
chart.type = 'col'
chart.title = 'Ventas Mensuales (USD)'
chart.y_axis.title = 'Total Ventas'
chart.x_axis.title = 'Mes'
chart.style = 10
data_ref = Reference(ws3, min_col=2, min_row=1, max_row=len(mensual)+1)
cats_ref = Reference(ws3, min_col=1, min_row=2, max_row=len(mensual)+1)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.width = 24
chart.height = 12
ws3.add_chart(chart, 'E5')

# ────── HOJA 4: Top productos por región ──────────────────
ws4 = wb.create_sheet('Por Region')
region_pivot = df.pivot_table(
    values='Total', index='Region', columns='Producto',
    aggfunc='sum', fill_value=0
)
region_pivot['Total Region'] = region_pivot.sum(axis=1)
for r in dataframe_to_rows(region_pivot.reset_index(), index=False, header=True):
    ws4.append(r)
for cell in ws4[1]:
    cell.font = header_font
    cell.fill = header_fill

# ────── HOJA 5: Resumen ejecutivo ─────────────────────────
ws5 = wb.create_sheet('Resumen')
resumen_data = [
    ['Métrica', 'Valor'],
    ['Total Ventas', f'${df["Total"].sum():,.2f}'],
    ['Transacciones', len(df)],
    ['Clientes únicos', df['Cliente'].nunique()],
    ['Productos únicos', df['Producto'].nunique()],
    ['Ticket promedio', f'${df["Total"].mean():,.2f}'],
    ['Período', f'{df["Fecha"].min().date()} → {df["Fecha"].max().date()}'],
    ['Mejor mes', mensual.loc[mensual['Total'].idxmax(), 'Mes']],
    ['Mejor cliente', df.groupby('Cliente')['Total'].sum().idxmax()],
    ['Producto más vendido', df.groupby('Producto')['Total'].sum().idxmax()],
]
for i, row in enumerate(resumen_data, 1):
    ws5.cell(row=i, column=1, value=row[0]).font = Font(bold=True)
    ws5.cell(row=i, column=2, value=row[1])

ws5.column_dimensions['A'].width = 22
ws5.column_dimensions['B'].width = 30

# Ajustar anchos
for ws in [ws1, ws2, ws3, ws4]:
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 25)

# ── Guardar ───────────────────────────────────────────────
wb.save(OUTPUT)
print(f'\n✓ Reporte Excel generado: {OUTPUT}')
print(f'  Hojas: Datos | Pivot Cliente-Producto | Ventas Mensuales | Por Region | Resumen')
